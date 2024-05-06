from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup


book_name="cert-validity.xlsx"
cert_col='A'
valid_col='B'  
start_row=2
wb=Workbook()
wb=load_workbook(book_name)
ws=wb.active
end_row = ws.max_row


def retrieve_current_until(cert_id):
    try:
        url = f'https://www.redhat.com/rhtapps/verify/?certId={cert_id}'
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            error_message = soup.find('p', class_='push-top alert alert-danger')
            if error_message:
                message_text = error_message.text.strip()
                if "not a valid Certification ID" in message_text:
                    return "invalid id"
                elif "not mapped his or her ID to a redhat.com login" in message_text:
                    return "not mapped yet"
            table = soup.find('table', {'style': 'width: 40%; white-space: nowrap'})
            if table:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    if len(cells) == 2 and cells[0].text.strip() == 'Current Until:':
                        return cells[1].text.strip()
                return "expired"
            else:
                return "expired"
        else:
            print(f"Error: {response.status_code} - Cannot able to load the page for ID {cert_id}")
            return "invalid"
    except Exception as e:
        print(f"Error: {e}")
        return "invalid"


#print(retrieve_current_until('230-223-091'))
   
for row in range(2, end_row+1):
    print(f'Executing {row-1} id')
    cell=f'{cert_col}{row}'
    cert_id=str(ws[cell].value)
    cert_id=cert_id.replace(' ','')
    valid_cur_row=f'{valid_col}{row}'
    if cert_id is None:
        ws[valid_cur_row]='invalid id'
        continue
    valid_date=retrieve_current_until(cert_id)
    ws[valid_cur_row]=valid_date

wb.save(book_name)
    

