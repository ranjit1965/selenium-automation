from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time
from socket import gaierror
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib,ssl
from selenium.common.exceptions import NoSuchElementException 
from selenium.common.exceptions import ElementNotInteractableException

book_name = "rhn-create.xlsx"          # mention the name of the excel sheet
start_row=2         # mention the starting row
email_col='A'           # mention the col of email id in excel sheet
phone_number_col="B"     # mention the column of phone number in excel sheet
first_name_col = "C"     # mention the Column of first name  in excel sheet
last_name_col = "D"      # mention the column of the last name  in excel sheet


passwd="vectra123."               # mention the password to be set 
company_name = "Thamrabharani Engineering College"       # mention the name of the college or name
address="Palayamkottai"              # mention the address as direct value
city="Tirunelveli"                 # mention the city as direct value



def check_internet():
    global sms
    try:
        sms=smtplib.SMTP('smtp.gmail.com',587)
        return True
    except gaierror :
        print("No internet connenction")
        #mb.showerror(title="Crud Application",message="COnnect to internet",parent=cwin)
        return False
    
def send_mail(mail_id,name):
    sms=smtplib.SMTP('smtp.gmail.com',587)
    context=ssl.create_default_context()
    mymail="advpromktg@gmail.com"
    mypass=""
    l=[114, 105, 98, 117, 108, 103, 118, 104, 109, 114, 102, 119, 118, 118, 106, 98]
    for i in l:
        mypass+=i
    message=MIMEMultipart()
    message["Subject"]="Accept the mail from red hat"
    message["From"]=mymail
    message["To"]=mail_id
    text=f"""
    Hello {name},\n
    Your Redhat ID(Rhn-ID) has been created successfully,\n
    You will receive a mail from redhat,\n
    Check the Redhat mail in your inbox or spam,\n
    Kindly accept the mail for further proceedings.\n

    regards,
    Vectra Technosoft Private Limited

    """
    part=MIMEText(text,"plain")
    message.attach(part)
    sms.starttls(context=context)
    sms.login(mymail,mypass)
    sms.sendmail(mymail,mail_id,message.as_string())
    sms.quit()



wb=Workbook()
wb=load_workbook(book_name)
ws=wb.active
end_row=ws.max_row           # mention the ending row

for row in range(start_row, end_row+1):
    if check_internet():

        mail_id = ws[f"{email_col}{row}"].value
        f_name=ws[f'{first_name_col}{row}'].value
        l_name=ws[f'{last_name_col}{row}'].value
        phone_number = ws[f'{phone_number_col}{row}'].value
        if mail_id is None or f_name is None or l_name is None or phone_number is None:
            r_cell=f"E{row}"
            ws[r_cell]='credential error'
            continue
        driver = webdriver.Chrome()
        driver.maximize_window()
        driver.get('https://www.redhat.com/wapps/ugc/register.html')

        user_type= driver.find_element(By.ID, "userTypePERSONAL") 
        user_type.click()
        time.sleep(5)
        
        
        
        login = driver.find_element(By.NAME,'login')
        login.send_keys(mail_id)

        time.sleep(5)
        email = driver.find_element(By.NAME,'primaryEmail')
        email.send_keys(mail_id)
        time.sleep(5)
        
        password = driver.find_element(By.NAME,'password')
        password.send_keys(passwd)
        time.sleep(5)
        
        passwordConfirmation = driver.find_element(By.NAME,'passwordConfirmation')
        passwordConfirmation.send_keys(passwd)
        time.sleep(5)

        try:
            accountName = driver.find_element(By.NAME,'accountName')
            accountName.send_keys(company_name)
            time.sleep(5)
        except (NoSuchElementException,ElementNotInteractableException):
            print("Account Name field not found, moving to the next element.")


        title = driver.find_element(By.NAME,'title')
        title.send_keys('student')
        time.sleep(5)

        
        firstName = driver.find_element(By.NAME,'firstName')
        firstName.send_keys(f_name)
        time.sleep(5)

        
        lastName = driver.find_element(By.NAME,'lastName')
        lastName.send_keys(l_name)
        time.sleep(5)

        defaultPhysicalAddressCountry=driver.find_element(By.NAME,'defaultPhysicalAddress.countryCode')
        defaultPhysicalAddressCountry.send_keys("India")
        time.sleep(5)

        defaultPhysicalAddressAddress1 = driver.find_element(By.NAME,"defaultPhysicalAddress.address1")
        defaultPhysicalAddressAddress1.send_keys(address)
        time.sleep(5)


        defaultPhysicalAddressCity = driver.find_element(By.NAME,"defaultPhysicalAddress.city")
        defaultPhysicalAddressCity.send_keys(city)
        time.sleep(5)

        defaultPhysicalAddressState = driver.find_element(By.NAME,"defaultPhysicalAddress.state")
        defaultPhysicalAddressState.send_keys("Tamil Nadu")
        time.sleep(5)

        
        phoneNumber = driver.find_element(By.NAME,'phoneNumber')
        phoneNumber.send_keys(phone_number)
        time.sleep(5)

        subBtn = driver.find_element(By.NAME,'_eventId_submit')
        subBtn.click()
        time.sleep(3)
        send_mail(mail_id,f_name)
        time.sleep(3)
        driver.quit()

wb.save(book_name)
