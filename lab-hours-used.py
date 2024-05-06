from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time



book_name="lab-hours-used.xlsx"       # mention name of the excel sheet
rhn_col="A"                 # mention the col of the rhn id in excel sheet
lab_usd_col="B"            # mention the col of the lab used in excel sheet
passwd="vectra123."          # mention the password of rhn id's
course_page_url='https://rha.ole.redhat.com/rha/app/courses/rh124-9.0/8609af49-75da-466f-a03c-ac5177405efb'

# mention the url of preferred class and course url in the course_page_url

wb=Workbook()
wb=load_workbook(book_name)
ws=wb.active
end_row = ws.max_row
print(end_row)
for row in range(2, end_row+1):
    
    rhn_cell = f"{rhn_col}{row}"
    data=ws[rhn_cell].value
    lab_used_cell=f"{lab_usd_col}{row}"
    if data is None:
        ws[lab_used_cell]="credential error"
        continue
    
    if data == 'No' or data == 'NO' or data == 'no':
        continue
    try:
        driver = webdriver.Chrome()

        #driver.maximize_window()
        
        driver.get("https://rha.ole.redhat.com/rha/app/login")

        time.sleep(15)
        email_elem = driver.find_element(By.NAME, "username") 
        email_elem.send_keys(data)
        cookie_btn = driver.find_element(By.ID,'truste-consent-button')
        cookie_btn.click()
        next_btn = driver.find_element(By.ID, "login-show-step2")  
        next_btn.click()
        
        
        
        time.sleep(5)
        password_elem = driver.find_element(By.ID, "password")
        password_elem.send_keys(passwd)
        
        time.sleep(5)
        
        login_box = driver.find_element(By.ID, 'rh-password-verification-submit-button')
        login_box.click()
        time.sleep(10)
        

        try:
            wait = WebDriverWait(driver, 7)
            element = wait.until(EC.visibility_of_element_located((By.ID, "rh-login-form-error-title")))
            if element is not  None :
                
                ws[lab_used_cell]="credential error"
                driver.quit()
                continue
        except TimeoutException:
            pass
        
        driver.get(course_page_url)
        
        """ #time.sleep(2)
        #driver.find_element(By.TAG_NAME,'body').send_keys(Keys.CONTROL, '-')
        driver.set_window_size(height=1080,width=1920)
        #action = ActionChains(driver)
        launch_box = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'launch-button') and contains(@class, 'ml-2') and contains(@class, 'btn') and contains(@class, 'btn-link')]")))
        #driver.execute_script("arguments[2].scrollIntoView(true);", launch_box)
        
        #action.move_to_element(launch_box).click()
        launch_box.click()
        time.sleep(5)

        """

        time.sleep(15)
        lab_box = driver.find_element(By.ID, 'course-tabs-tab-8')
        lab_box.click()
        time.sleep(7) 
        
        lab_hour = driver.find_element(By.CLASS_NAME, 'instruction-wrapper')
        
        used=str(lab_hour.text)
        lab_used_cell=f"{lab_usd_col}{row}"
        ws[lab_used_cell]=used[16:]
        
        driver.quit()
    except Exception as e:
        print(e)
        wb.save(book_name)


wb.save(book_name)
    

